# -*- coding: utf-8 -*-
import os
import re
import sys
import threading
import tkinter as tk
from collections import defaultdict
from tkinter import ttk, filedialog, messagebox

import openpyxl
from openpyxl.utils import get_column_letter


class SpeciesProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("物种数据整理工具")
        self.root.geometry("800x600")
        self.setup_ui()
        self.running = False  # 添加运行状态标志

    def setup_ui(self):
        # 创建主框架
        main_frame = ttk.Frame(self.root, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 标题
        title_label = ttk.Label(
            main_frame,
            text="物种样地数据整理工具",
            font=("Arial", 14, "bold")
        )
        title_label.pack(pady=10)

        # 说明文本
        description = tk.Text(
            main_frame,
            height=5,
            width=80,
            font=("Arial", 9),
            bg="#f0f0f0",
            padx=10,
            pady=10
        )
        description.insert(tk.END, "使用说明：\n")
        description.insert(tk.END, "1. 输入文件应为Excel格式，包含原始物种数据\n")
        description.insert(tk.END, "2. 数据格式要求：每个样地以'物种名称\\t样地编号'开头，后跟物种列表\n")
        description.insert(tk.END, "3. 程序会自动合并相同物种的数量，并生成物种-样地矩阵\n")
        description.insert(tk.END, "4. 植物名称和代码在同一列，样地为列标题\n")
        description.configure(state=tk.DISABLED)
        description.pack(pady=10, fill=tk.X)

        # 输入文件选择
        input_frame = ttk.Frame(main_frame)
        input_frame.pack(fill=tk.X, pady=10)

        ttk.Label(input_frame, text="选择输入文件:").pack(side=tk.LEFT, padx=(0, 10))
        self.input_entry = ttk.Entry(input_frame, width=50)
        self.input_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))

        ttk.Button(
            input_frame,
            text="浏览...",
            command=self.browse_input_file
        ).pack(side=tk.RIGHT)

        # 输出文件选择
        output_frame = ttk.Frame(main_frame)
        output_frame.pack(fill=tk.X, pady=10)

        ttk.Label(output_frame, text="输出文件路径:").pack(side=tk.LEFT, padx=(0, 10))
        self.output_entry = ttk.Entry(output_frame, width=50)
        self.output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))

        ttk.Button(
            output_frame,
            text="浏览...",
            command=self.browse_output_file
        ).pack(side=tk.RIGHT)

        # 处理按钮
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=20)

        self.process_btn = ttk.Button(
            button_frame,
            text="开始处理数据",
            command=self.process_data,
            width=20
        )
        self.process_btn.pack()

        # 添加停止按钮
        self.stop_btn = ttk.Button(
            button_frame,
            text="停止处理",
            command=self.stop_processing,
            width=20,
            state=tk.DISABLED
        )
        self.stop_btn.pack(pady=5)

        # 进度条
        self.progress = ttk.Progressbar(
            main_frame,
            orient=tk.HORIZONTAL,
            length=400,
            mode='determinate'
        )
        self.progress.pack(pady=10)

        # 进度标签
        self.progress_label = ttk.Label(main_frame, text="就绪")
        self.progress_label.pack(pady=5)

        # 日志输出
        log_frame = ttk.LabelFrame(main_frame, text="处理日志")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        self.log_text = tk.Text(
            log_frame,
            height=10,
            bg="#f8f8f8",
            state=tk.DISABLED,
            font=("Arial", 9)
        )
        scrollbar = ttk.Scrollbar(log_frame, command=self.log_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 状态栏
        self.status_var = tk.StringVar()
        self.status_var.set("就绪")
        status_bar = ttk.Label(
            main_frame,
            textvariable=self.status_var,
            relief=tk.SUNKEN,
            anchor=tk.W
        )
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)

        # 设置默认输出路径
        self.output_entry.insert(0, os.path.join(os.getcwd(), "物种样地矩阵.xlsx"))

    def browse_input_file(self):
        file_path = filedialog.askopenfilename(
            title="选择输入文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if file_path:
            self.input_entry.delete(0, tk.END)
            self.input_entry.insert(0, file_path)

            # 自动生成输出文件名
            dir_name, file_name = os.path.split(file_path)
            base_name = os.path.splitext(file_name)[0]
            output_path = os.path.join(dir_name, f"{base_name}_矩阵.xlsx")
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, output_path)

    def browse_output_file(self):
        file_path = filedialog.asksaveasfilename(
            title="保存输出文件",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
            defaultextension=".xlsx"
        )
        if file_path:
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, file_path)

    def log_message(self, message):
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state=tk.DISABLED)
        self.root.update()

    def update_progress(self, value, message=None):
        self.progress['value'] = value
        if message:
            self.progress_label.config(text=message)
        self.root.update()

    def stop_processing(self):
        self.running = False
        self.log_message("处理已停止...")
        self.status_var.set("已停止")
        self.process_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)

    def process_data(self):
        if self.running:
            return

        self.running = True
        input_path = self.input_entry.get()
        output_path = self.output_entry.get()

        if not input_path:
            messagebox.showerror("错误", "请选择输入文件")
            self.running = False
            return

        if not output_path:
            messagebox.showerror("错误", "请指定输出文件路径")
            self.running = False
            return

        try:
            self.process_btn.config(state=tk.DISABLED)
            self.stop_btn.config(state=tk.NORMAL)
            self.log_text.configure(state=tk.NORMAL)
            self.log_text.delete(1.0, tk.END)
            self.log_text.configure(state=tk.DISABLED)

            self.log_message("=" * 50)
            self.log_message(f"开始处理数据: {os.path.basename(input_path)}")
            self.status_var.set("处理中...")

            # 在后台线程中处理数据
            thread = threading.Thread(
                target=self.process_data_thread,
                args=(input_path, output_path),
                daemon=True
            )
            thread.start()

        except Exception as e:
            self.log_message(f"错误: {str(e)}")
            messagebox.showerror("处理错误", f"处理过程中发生错误:\n{str(e)}")
            self.status_var.set("处理失败")
            self.process_btn.config(state=tk.NORMAL)
            self.stop_btn.config(state=tk.DISABLED)
            self.running = False

    def process_data_thread(self, input_path, output_path):
        try:
            # 读取原始数据
            self.log_message("读取Excel文件...")
            self.update_progress(5, "正在读取文件...")

            # 使用openpyxl读取Excel
            wb = openpyxl.load_workbook(input_path, data_only=True)
            sheet = wb.active

            # 数据预处理
            self.log_message("解析数据...")
            self.update_progress(10, "正在解析数据...")
            current_plot = None
            plot_data = defaultdict(dict)  # 存储每个样地的物种数据
            all_species = set()  # 所有唯一物种集合
            plot_counter = 0  # 样地计数器
            species_counter = 0  # 物种记录计数器

            total_rows = sheet.max_row
            processed_rows = 0
            last_progress = 0

            # 样地识别模式
            plot_patterns = [
                r'[0-9]+-[0-9]+-[0-9]+',  # 1-1-1 格式
                r'[0-9]+-[0-9]+',  # 1-1 格式
                r'[0-9]+'  # 1 格式
            ]

            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                if not self.running:
                    self.log_message("处理已中断")
                    return

                # 跳过空行
                if not row or all(cell is None for cell in row):
                    processed_rows += 1
                    continue

                # 更新进度（每10行或10%更新一次）
                processed_rows += 1
                current_progress = 10 + int(70 * processed_rows / total_rows)
                if current_progress > last_progress + 5 or row_idx == total_rows:
                    self.update_progress(
                        current_progress,
                        f"处理中: {processed_rows}/{total_rows} 行"
                    )
                    last_progress = current_progress

                # 检查是否为样地行
                is_plot_row = False
                if row[0] and isinstance(row[0], str) and "物种名称" in row[0]:
                    # 尝试匹配多种样地编号格式
                    for pattern in plot_patterns:
                        plot_match = re.search(pattern, str(row[0]))
                        if plot_match:
                            current_plot = plot_match.group()
                            self.log_message(f"发现样地: {current_plot}")
                            plot_counter += 1
                            is_plot_row = True
                            break

                    # 如果第一列包含"物种名称"关键词
                    if not is_plot_row and len(row) > 1 and row[1]:
                        # 尝试从第二列获取样地编号
                        current_plot = str(row[1]).strip()
                        self.log_message(f"发现样地: {current_plot}")
                        plot_counter += 1
                        is_plot_row = True

                if is_plot_row:
                    continue

                # 处理物种数据行
                if current_plot and row[0] and isinstance(row[0], str):
                    species_name = str(row[0]).strip()

                    if not species_name or "物种名称" in species_name:
                        continue

                    # 获取数量值
                    count = 0
                    if len(row) > 1 and row[1] is not None:
                        # 尝试直接转换为数字
                        if isinstance(row[1], (int, float)):
                            count = row[1]
                        else:
                            try:
                                # 尝试转换字符串为数字
                                count = float(str(row[1]))
                            except (ValueError, TypeError):
                                # 从字符串中提取数字
                                num_str = re.search(r'[\d.]+', str(row[1]))
                                if num_str:
                                    try:
                                        count = float(num_str.group())
                                    except:
                                        count = 0
                                else:
                                    count = 0
                    elif len(row) > 1 and row[1] is not None and isinstance(row[1], (int, float)):
                        # 如果第二列是数字，可能是数量值
                        count = row[1]

                    # 统计原始数据行数
                    species_counter += 1

                    # 累加相同物种的数量
                    if species_name in plot_data[current_plot]:
                        # 如果物种已存在，累加数量
                        plot_data[current_plot][species_name] += count
                        self.log_message(
                            f"  累加物种: {species_name} + {count} = {plot_data[current_plot][species_name]}")
                    else:
                        # 如果物种不存在，添加新记录
                        plot_data[current_plot][species_name] = count
                        self.log_message(f"  添加物种: {species_name} = {count}")

                    # 添加到总物种集合
                    all_species.add(species_name)

            if not current_plot:
                messagebox.showerror("错误", "未找到样地数据！请检查文件格式")
                return

            # 创建唯一物种列表并排序
            unique_species = sorted(all_species)
            self.log_message(f"发现 {len(unique_species)} 个唯一物种")
            self.log_message(f"发现 {plot_counter} 个样地")
            self.log_message(f"处理了 {species_counter} 条物种记录")

            # 创建矩阵数据结构
            self.log_message("创建物种-样地矩阵...")
            self.update_progress(85, "正在生成矩阵...")

            # 对样地排序（按数字顺序）
            def plot_key(plot):
                parts = plot.split('-')
                try:
                    return tuple(int(part) for part in parts)
                except ValueError:
                    return tuple(part for part in parts)

            sorted_plots = sorted(plot_data.keys(), key=plot_key)

            # 创建新的工作簿
            output_wb = openpyxl.Workbook()
            output_sheet = output_wb.active
            output_sheet.title = "物种样地矩阵"

            # 写入表头（样地为列）
            header = ['物种'] + sorted_plots
            output_sheet.append(header)

            # 写入数据（物种为行）
            for species in unique_species:
                if not self.running:
                    self.log_message("处理已中断")
                    return

                row = [species]
                for plot in sorted_plots:
                    # 获取该物种在该样地的总数量
                    count = plot_data[plot].get(species, 0)
                    row.append(count)
                output_sheet.append(row)

            # 自动调整列宽
            self.log_message("优化表格格式...")
            for col_idx, _ in enumerate(header, 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                for row in output_sheet.iter_rows(min_row=1, max_row=output_sheet.max_row, min_col=col_idx,
                                                  max_col=col_idx):
                    for cell in row:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                adjusted_width = (max_length + 2) * 1.2
                output_sheet.column_dimensions[col_letter].width = adjusted_width

            # 保存结果
            self.log_message(f"保存结果到: {output_path}")
            self.update_progress(95, "正在保存文件...")
            output_wb.save(output_path)

            # 完成
            self.log_message("数据处理完成!")
            self.log_message(f"包含物种数量: {len(unique_species)}")
            self.log_message(f"包含样地数量: {len(plot_data)}")
            self.log_message(f"原始物种记录数: {species_counter}")
            self.update_progress(100, "处理完成!")

            messagebox.showinfo("完成", f"数据处理完成!\n结果已保存到: {output_path}")
            self.status_var.set("处理完成")

        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            self.log_message(f"处理过程中发生错误: {str(e)}")
            self.log_message("详细错误信息:")
            self.log_text.configure(state=tk.NORMAL)
            self.log_text.insert(tk.END, error_trace)
            self.log_text.configure(state=tk.DISABLED)
            self.log_text.see(tk.END)

            messagebox.showerror("处理错误", f"处理过程中发生错误:\n{str(e)}")
            self.status_var.set("处理失败")

        finally:
            self.process_btn.config(state=tk.NORMAL)
            self.stop_btn.config(state=tk.DISABLED)
            self.running = False


if __name__ == "__main__":
    # 创建主窗口
    root = tk.Tk()

    # 设置DPI感知（Windows）
    if sys.platform == "win32":
        try:
            from ctypes import windll

            windll.shcore.SetProcessDpiAwareness(1)
        except:
            pass

    # 创建应用实例
    app = SpeciesProcessorApp(root)

    # 启动主循环
    root.mainloop()